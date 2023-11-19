"""
This app is described in the readme.md
current version 1.0.1
"""


import math
import os
import itertools
import glob
from openpyxl import load_workbook

example_file = "random_values.xlsx"
report_path = "reports/"
isexample = False
banner = "*" * 40 + "\n"

header = [banner, "WELCOME TO THE PRICE IS RIGHT\n", banner]
CELLINDEX, SHEETINDEX, DOCINDEX = 0, 1, 2
data_loc = ["", "", "", banner]

POPIDX, POSIDX, CONIDX, COMBIDX = 1, 2, 3, 4
jobsum = [banner, "", "", "", "", banner]

population = []
combo_status = {}
found = []


def total_combinations(n):
    # total = 0
    # for k in range(2, n + 1):
    #     total += math.comb(n, k)
    # return total

    totals = {}
    for k in range(2, n + 1):
        totals[k] = math.comb(n, k)

    return sum(totals.values()), totals


def printStatus(tofile=False):
    rep_lines = []

    rep_lines.append("\nCombos of".rjust(10))
    rep_lines[0] += "Possible".rjust(10)
    rep_lines[0] += "Completed".rjust(12)
    rep_lines[0] += "  Found\n"

    for key, vals in combos_reports.items():
        # combos_reports[c_size] = [total, searches, failures, good_combos]
        a, b, c, d = vals
        thisrow = ""
        thisrow += str(key).rjust(9)
        thisrow += str(a).rjust(10)
        thisrow += str(b).rjust(12)
        thisrow += f"  {len(found)}"
        thisrow += "\n"
        rep_lines.append(thisrow)

    rep_lines.append(f"\nCombo found: {found}")

    status_msg = "".join(rep_lines)

    print(status_msg)

    if tofile:
        return status_msg
    return False


def printScreen(tofile=False):
    reptext = "".join(header) + "\n"
    
    reptext += "\n".join(data_loc)
    reptext += "\n".join(jobsum)
    os.system("cls")
    print(reptext)

    if tofile:
        return reptext
    return False


# PART 1: OPEN THE EXEL FILE AND PULL OUT OUR POPULATION DATA
# Load the Excel file
printScreen()
types = ("*.xlsx", "*.xls")  # the tuple of file types
files_grabbed = []
for files in types:
    files_grabbed.extend(glob.glob(files))

available_spreadsheets = {id + 1: x for id, x in enumerate(files_grabbed)}
query = "What is the name of the DOCUMENT file"
query += " we need to search?\n##  Filename\n"

for i, f in available_spreadsheets.items():
    query += f"{i})  {f}\n"

query += f"\nSelect file from index 1 - {len(available_spreadsheets)} > "

response = input(query)

isexample = response == ""
if isexample:
    # default for testing example
    filename = example_file
else:
    filename = available_spreadsheets[int(response)]

wb = load_workbook(filename, data_only=True)

data_loc[DOCINDEX] = "Searching DOCUMENT:".rjust(18) + f" '{filename}'."

# Select the worksheet
printScreen()
query = "What is the SHEETNAME within the "
query += "DOCUMENT file we need to search?\n## Sheetname"

list_of_sheets = {i + 1: sh for i, sh in enumerate(wb.sheetnames)}
# print list of available sheets to screen

for i, x in list_of_sheets.items():
    query += f"\n{i})  {x}"

response = input(query + "\nDefault is #1> ")
if not response.isnumeric():
    # default for testing example
    sheetname = list_of_sheets[1]
else:
    response = int(response)
    sheetname = list_of_sheets[response]

data_loc[SHEETINDEX] = "From SHEETNAME:".rjust(18) + f" '{sheetname}'"
ws = wb[sheetname]

# Define the range of cells to read
printScreen()
query = "What is the FIRST CELL of the range we need to search?\n"
query += "Default is 'C2'  > "
start_cell = input(query)

query = "\nWhat is the LAST CELL of the range we need to search?\n"
query += "Default is 'C22' > "
end_cell = input(query)

if not start_cell.isalnum():
    start_cell = "C2"

if not end_cell.isalnum():
    # default for testing example
    end_cell = "C22"


data_loc[CELLINDEX] = "Searching RANGE:".rjust(18)
data_loc[CELLINDEX] += f" '{start_cell}:{end_cell}'"

printScreen()
# Read the cell values into a list of lists
for row in ws[start_cell:end_cell]:
    for cell in row:
        if cell.value is not None:
            population.append(cell.value)
population_str = ["$ {:,}".format(x) for x in population]


pop_count = len(population)

t_possibilities = 0
combo_possibilities = {}
for n in range(2, pop_count + 1):
    # Generate all possible combinations of combo_size elements
    combos_list = list(itertools.combinations(population, n))
    combo_possibilities[n] = combos_list
    t_possibilities += len(combos_list)

combinations_str = "{:,}".format(t_possibilities)

# job summary
jobsum[POPIDX] = f"Population Total: {pop_count}"
jobsum[POSIDX] = f"Total Possible Combinations: {combinations_str}"
jobsum[CONIDX] = f"Contents: \n{', '.join(population_str)}\n"


printScreen()

# get the target value we are trying to find
targetfromuser = input(
    "What is the TARGET VALUE we are trying "
    + "to find?\n(no dollar signs or commas please.)\n> "
)
if targetfromuser.isnumeric():
    target_sum = float(targetfromuser)
else:
    target_sum = sum([870776.16, 948031.4, 811827.41])

jobsum[COMBIDX] = "Target value combinations must sum: $"
jobsum[COMBIDX] += " {:,}".format(target_sum)
printScreen()

_ = input("Press any key to continue.\n> ")


# first I'm gonna try to brute force it.

# combos_reports = {0 : [total, searched, failures, [successes]]}
combos_reports = {}


def is_match(combo):
    truecondition = sum(combo) == target_sum
    if truecondition:
        vals = ", ".join(["$ {:,.2f}".format(x) for x in combo])
        found.append(vals)
    return truecondition


def dothiscombo(c_size, c_list):
    total = len(c_list)
    searches = 0
    failures = 0
    good_combos = []

    for c in c_list:
        searches += 1
        if is_match(c):
            good_combos.append(c)

        else:
            failures += 1

        combos_reports[c_size] = [total, searches, failures, good_combos]


for combo_size, combinations_list in combo_possibilities.items():
    dothiscombo(combo_size, combinations_list)
    printScreen()
    printStatus()


def save_file(reportname):
    try:
        text = printScreen(tofile=True) + printStatus(tofile=True)
        open(reportname, "w").write(text)

    except Exception:
        print("ERROR: PROBLEM MAKING THE FILE!")

        try:
            print("ATTEMPT: Trying to make the folder.")
            os.mkdir(report_path)
        except Exception:
            print(f"ERROR: Can't fucking make a dir called {report_path}")

            open(reportname, "w").write(
                printScreen(tofile=True) + printStatus(tofile=True)
            )


report_count = str(len(list(glob.glob("reports/*.txt"))) + 1).rjust(3, "0")

save_file_name = report_path
if isexample:
    save_file_name += f"report_({report_count})_{example_file[:-5]}.txt"
else:
    save_file_name += f"report_({report_count})_{filename[:-5]}.txt"

print(f"Saving File: {save_file_name}")
save_file(save_file_name)
